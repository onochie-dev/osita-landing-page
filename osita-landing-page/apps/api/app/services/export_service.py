"""
Export Service
Generates Excel and XML exports from canonical data.
"""
import io
import os
import zipfile
from typing import Dict, Any, Optional, List
from datetime import datetime
from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from lxml import etree

from ..config import get_settings


class ExportService:
    """
    Service for generating exports in various formats.
    """
    
    # CBAM XML namespace (based on official schema)
    CBAM_NS = "urn:cbam:report:v1"
    
    def __init__(self):
        self.settings = get_settings()
    
    def generate_excel(
        self,
        project_data: Dict[str, Any],
        canonical_data: Dict[str, Any],
        validation_flags: List[Dict[str, Any]],
        evidence_items: List[Dict[str, Any]]
    ) -> bytes:
        """
        Generate Excel export with multiple sheets.
        
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary_sheet(ws_summary, project_data, canonical_data, header_font, header_fill, border)
        
        # Bills Sheet
        ws_bills = wb.create_sheet("Electricity Bills")
        self._write_bills_sheet(ws_bills, canonical_data, header_font, header_fill, border)
        
        # Emissions Sheet
        ws_emissions = wb.create_sheet("Indirect Emissions")
        self._write_emissions_sheet(ws_emissions, canonical_data, header_font, header_fill, border)
        
        # Validation Sheet
        ws_validation = wb.create_sheet("Validation Summary")
        self._write_validation_sheet(ws_validation, validation_flags, header_font, header_fill, border)
        
        # Evidence Sheet
        ws_evidence = wb.create_sheet("Evidence Trail")
        self._write_evidence_sheet(ws_evidence, evidence_items, header_font, header_fill, border)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
    
    def _write_summary_sheet(self, ws, project_data, canonical_data, header_font, header_fill, border):
        """Write summary sheet."""
        # Title
        ws["A1"] = "CBAM Quarterly Report - Indirect Emissions (Electricity)"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:D1")
        
        # Project Info
        rows = [
            ("Project Name", project_data.get("name", "")),
            ("Reporting Period", f"{canonical_data.get('reporting_period', 'N/A')} {canonical_data.get('reporting_year', '')}"),
            ("Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
            ("", ""),
            ("Total Electricity Consumed (MWh)", canonical_data.get("total_electricity_mwh", 0)),
            ("Total Indirect Emissions (tCO₂)", canonical_data.get("total_indirect_emissions_tco2", 0)),
            ("Emission Factor Source", project_data.get("emission_factor_source", "Commission default")),
            ("Emission Factor Value (tCO₂/MWh)", project_data.get("emission_factor_value", "N/A")),
        ]
        
        for idx, (label, value) in enumerate(rows, start=3):
            ws[f"A{idx}"] = label
            ws[f"A{idx}"].font = Font(bold=True)
            ws[f"B{idx}"] = value
        
        # Declarant Info
        declarant = project_data.get("declarant_info", {})
        if declarant:
            start_row = len(rows) + 5
            ws[f"A{start_row}"] = "Declarant Information"
            ws[f"A{start_row}"].font = Font(bold=True, size=12)
            
            decl_rows = [
                ("Name", declarant.get("name", "")),
                ("Identification Number", declarant.get("identification_number", "")),
                ("Role", declarant.get("role", "")),
            ]
            
            for idx, (label, value) in enumerate(decl_rows, start=start_row + 1):
                ws[f"A{idx}"] = label
                ws[f"B{idx}"] = value
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 40
    
    def _write_bills_sheet(self, ws, canonical_data, header_font, header_fill, border):
        """Write electricity bills sheet."""
        headers = ["Document ID", "Supplier", "Period Start", "Period End", 
                   "Consumption", "Unit", "Normalized (MWh)", "Amount", "Currency"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        bills = canonical_data.get("electricity_bills", [])
        for row_idx, bill in enumerate(bills, start=2):
            tc = bill.get("total_consumption", {})
            bp = bill.get("billing_period", {})
            
            values = [
                bill.get("document_id", ""),
                bill.get("supplier", ""),
                bp.get("start_date", ""),
                bp.get("end_date", ""),
                tc.get("value", ""),
                tc.get("unit", ""),
                tc.get("normalized_mwh", ""),
                bill.get("total_amount", ""),
                bill.get("currency", "")
            ]
            
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
        
        # Adjust widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _write_emissions_sheet(self, ws, canonical_data, header_font, header_fill, border):
        """Write indirect emissions sheet."""
        headers = ["Period Start", "Period End", "Electricity (MWh)", 
                   "Emission Factor", "Factor Source", "Emissions (tCO₂)"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        emissions = canonical_data.get("indirect_emissions", [])
        for row_idx, em in enumerate(emissions, start=2):
            values = [
                em.get("period_start", ""),
                em.get("period_end", ""),
                em.get("electricity_consumed_mwh", 0),
                em.get("emission_factor", ""),
                em.get("emission_factor_source", ""),
                em.get("emissions_tco2", 0)
            ]
            
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
        
        # Total row
        if emissions:
            total_row = len(emissions) + 2
            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=total_row, column=3, value=canonical_data.get("total_electricity_mwh", 0))
            ws.cell(row=total_row, column=6, value=canonical_data.get("total_indirect_emissions_tco2", 0))
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _write_validation_sheet(self, ws, flags, header_font, header_fill, border):
        """Write validation summary sheet."""
        headers = ["Code", "Severity", "Category", "Message", "Suggestion", "Status"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        severity_colors = {
            "blocking": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            "warning": PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
            "info": PatternFill(start_color="4ECDC4", end_color="4ECDC4", fill_type="solid")
        }
        
        for row_idx, flag in enumerate(flags, start=2):
            values = [
                flag.get("code", ""),
                flag.get("severity", ""),
                flag.get("category", ""),
                flag.get("message", ""),
                flag.get("suggestion", ""),
                "Resolved" if flag.get("is_resolved") else ("Acknowledged" if flag.get("is_acknowledged") else "Open")
            ]
            
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if col == 2:  # Severity column
                    cell.fill = severity_colors.get(value, PatternFill())
        
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 40
        ws.column_dimensions["F"].width = 15
    
    def _write_evidence_sheet(self, ws, evidence_items, header_font, header_fill, border):
        """Write evidence trail sheet."""
        headers = ["Field", "Value", "Document", "Page", "Quote", "Confidence"]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(evidence_items, start=2):
            values = [
                item.get("field_name", ""),
                item.get("value", ""),
                item.get("document_id", ""),
                item.get("source_page", ""),
                item.get("source_quote", ""),
                f"{item.get('confidence', 0):.0%}" if item.get("confidence") else ""
            ]
            
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
        
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 60
        ws.column_dimensions["F"].width = 12
    
    def generate_xml(
        self,
        project_data: Dict[str, Any],
        canonical_data: Dict[str, Any]
    ) -> str:
        """
        Generate CBAM-compliant XML.
        
        Returns:
            XML string
        """
        # Create root element
        root = etree.Element("Qreport", nsmap={None: self.CBAM_NS})
        
        # Reporting Period
        etree.SubElement(root, "ReportingPeriod").text = canonical_data.get("reporting_period", "Q1")
        etree.SubElement(root, "Year").text = canonical_data.get("reporting_year", str(datetime.now().year))
        
        # Totals
        etree.SubElement(root, "TotalImported").text = str(canonical_data.get("total_electricity_mwh", 0))
        etree.SubElement(root, "TotalEmissions").text = f"{canonical_data.get('total_indirect_emissions_tco2', 0):.7f}"
        
        # Declarant Type
        declarant_info = project_data.get("declarant_info", {})
        if declarant_info:
            declarant = etree.SubElement(root, "DeclarantType")
            etree.SubElement(declarant, "IdentificationNumber").text = declarant_info.get("identification_number", "")
            etree.SubElement(declarant, "Name").text = declarant_info.get("name", "")
            if declarant_info.get("role"):
                etree.SubElement(declarant, "Role").text = declarant_info["role"]
            
            # Actor Address
            address = declarant_info.get("address", {})
            if address:
                addr_elem = etree.SubElement(declarant, "ActorAddress")
                if address.get("country"):
                    etree.SubElement(addr_elem, "Country").text = address["country"]
                if address.get("sub_division"):
                    etree.SubElement(addr_elem, "SubDivision").text = address["sub_division"]
                if address.get("city"):
                    etree.SubElement(addr_elem, "City").text = address["city"]
                if address.get("street"):
                    etree.SubElement(addr_elem, "Street").text = address["street"]
                if address.get("street_additional_line"):
                    etree.SubElement(addr_elem, "StreetAdditionalLine").text = address["street_additional_line"]
                if address.get("number"):
                    etree.SubElement(addr_elem, "Number").text = address["number"]
                if address.get("postcode"):
                    etree.SubElement(addr_elem, "Postcode").text = address["postcode"]
                if address.get("po_box"):
                    etree.SubElement(addr_elem, "POBox").text = address["po_box"]
        
        # Imported Goods (Electricity section)
        imported_good = etree.SubElement(root, "ImportedGood")
        etree.SubElement(imported_good, "ItemNumber").text = "1"
        etree.SubElement(imported_good, "ImportArea").text = "EU"
        
        # Commodity code for electricity (CN code 2716)
        etree.SubElement(imported_good, "CNCommodityCode").text = "27160000"
        
        # Electricity consumption
        etree.SubElement(imported_good, "GoodMeasure").text = str(canonical_data.get("total_electricity_mwh", 0))
        etree.SubElement(imported_good, "GoodMeasureUnit").text = "MWH"
        
        # Indirect Emissions
        indirect = etree.SubElement(imported_good, "IndirectEmissions")
        emission_factor = project_data.get("emission_factor_value") or "0.4"  # Default factor
        etree.SubElement(indirect, "ElectricityConsumed").text = str(canonical_data.get("total_electricity_mwh", 0))
        etree.SubElement(indirect, "ElectricityEmissionFactor").text = str(emission_factor)
        etree.SubElement(indirect, "IndirectEmissionsValue").text = f"{canonical_data.get('total_indirect_emissions_tco2', 0):.7f}"
        etree.SubElement(indirect, "EmissionFactorSource").text = project_data.get("emission_factor_source", "DefaultValue")
        
        # Generate XML string with pretty printing
        return etree.tostring(
            root,
            pretty_print=True,
            xml_declaration=True,
            encoding="UTF-8"
        ).decode("utf-8")
    
    def generate_zip(
        self,
        project_data: Dict[str, Any],
        canonical_data: Dict[str, Any],
        validation_flags: List[Dict[str, Any]],
        evidence_items: List[Dict[str, Any]]
    ) -> bytes:
        """
        Generate ZIP package with XML and Excel.
        
        Returns:
            ZIP file as bytes
        """
        output = io.BytesIO()
        
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Add XML
            xml_content = self.generate_xml(project_data, canonical_data)
            zf.writestr("cbam_report.xml", xml_content)
            
            # Add Excel
            excel_content = self.generate_excel(
                project_data, canonical_data, validation_flags, evidence_items
            )
            zf.writestr("cbam_report.xlsx", excel_content)
            
            # Add metadata JSON
            metadata = {
                "generated_at": datetime.utcnow().isoformat(),
                "project_name": project_data.get("name"),
                "reporting_period": f"{canonical_data.get('reporting_period')} {canonical_data.get('reporting_year')}",
                "schema_version": "17.03"
            }
            zf.writestr("metadata.json", json.dumps(metadata, indent=2))
        
        output.seek(0)
        return output.read()

